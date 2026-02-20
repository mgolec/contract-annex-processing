"""Phase 1 output: Generate control spreadsheet (output/control_spreadsheet.xlsx)."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from doc_pipeline.config import PipelineConfig
from doc_pipeline.models import (
    ClientExtraction,
    ConfidenceLevel,
    Currency,
    Inventory,
)

# ── Style constants ──────────────────────────────────────────────────────────

_HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_LOCKED_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
_EDITABLE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
_BODY_FONT = Font(name="Arial", size=10)
_THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

# Conditional formatting fills for status column
_GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_BLUE_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

# Protection
_LOCKED = Protection(locked=True)
_UNLOCKED = Protection(locked=False)

# HRK conversion rate — removed hardcoded constant; now passed from config


# ── Confidence label mapping ─────────────────────────────────────────────────

_CONFIDENCE_LABELS = {
    ConfidenceLevel.HIGH: "Visoka",
    ConfidenceLevel.MEDIUM: "Srednja",
    ConfidenceLevel.LOW: "Niska",
}


# ── Helper: style a header row ───────────────────────────────────────────────


def _style_header(ws: Worksheet, headers: list[str], widths: list[int]) -> None:
    """Write header row and set column widths."""
    for col_idx, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER
        cell.protection = _LOCKED
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _style_cell(ws: Worksheet, row: int, col: int, value, *, locked: bool = True) -> None:
    """Write and style a body cell."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = _BODY_FONT
    cell.fill = _LOCKED_FILL if locked else _EDITABLE_FILL
    cell.border = _THIN_BORDER
    cell.protection = _LOCKED if locked else _UNLOCKED
    cell.alignment = Alignment(vertical="center", wrap_text=True)


# ── Sheet 1: Pregled klijenata ───────────────────────────────────────────────


def _build_sheet1(
    ws: Worksheet,
    extractions: list[ClientExtraction],
    inventory: Inventory,
) -> None:
    """Build 'Pregled klijenata' (Client Overview) sheet."""
    ws.title = "Pregled klijenata"

    headers = [
        "Klijent",
        "Mapa",
        "Glavni dokument",
        "Datum ugovora",
        "Posljednji aneks",
        "Datum aneksa",
        "Referentni dokument",
        "Pouzdanost",
        "Status",
        "Napomene",
        "Datum pregleda",
    ]
    widths = [25, 20, 30, 15, 30, 15, 30, 14, 16, 30, 15]
    _style_header(ws, headers, widths)

    # Build a lookup from folder name to inventory client
    inv_map = {c.folder_name: c for c in inventory.clients}

    # Sort extractions by folder name
    sorted_ext = sorted(extractions, key=lambda e: e.folder_name.lower())

    for row_idx, ce in enumerate(sorted_ext, 2):
        ex = ce.extraction
        inv_client = inv_map.get(ce.folder_name)
        chain = inv_client.document_chain if inv_client else None

        # Column A: Client name
        client_name = ex.client_name if ex and ex.client_name else ce.folder_name
        _style_cell(ws, row_idx, 1, client_name)

        # Column B: Folder
        _style_cell(ws, row_idx, 2, ce.folder_name)

        # Column C: Main document
        main_doc = ""
        if chain and chain.main_contract:
            main_doc = Path(chain.main_contract).name
        _style_cell(ws, row_idx, 3, main_doc)

        # Column D: Contract date
        doc_date = ""
        if ex and ex.document_type == "contract" and ex.document_date:
            doc_date = ex.document_date
        _style_cell(ws, row_idx, 4, doc_date)

        # Column E: Latest annex
        latest_annex = ""
        if chain and chain.annexes:
            latest_annex = Path(chain.annexes[-1]).name
        _style_cell(ws, row_idx, 5, latest_annex or "—")

        # Column F: Annex date
        annex_date = ""
        if ex and ex.document_type == "annex" and ex.document_date:
            annex_date = ex.document_date
        _style_cell(ws, row_idx, 6, annex_date or "—")

        # Column G: Reference document (source of extraction)
        ref_doc = Path(ce.source_file).name if ce.source_file else ""
        _style_cell(ws, row_idx, 7, ref_doc)

        # Column H: Confidence
        confidence_label = ""
        if ex:
            confidence_label = _CONFIDENCE_LABELS.get(ex.confidence, "")
        _style_cell(ws, row_idx, 8, confidence_label)

        # Column I: Status (editable)
        _style_cell(ws, row_idx, 9, "", locked=False)

        # Column J: Notes (editable)
        notes_text = ""
        if ce.error:
            notes_text = f"GREŠKA: {ce.error}"
        elif ex and ex.notes:
            notes_text = "; ".join(ex.notes)
        _style_cell(ws, row_idx, 10, notes_text, locked=False)

        # Column K: Review date (editable)
        _style_cell(ws, row_idx, 11, "", locked=False)

    # Data validation for Status column (I)
    last_row = len(sorted_ext) + 1
    dv = DataValidation(
        type="list",
        formula1='"Odobreno,Odbijeno,Preskočeno,Za raspravu"',
        allow_blank=True,
    )
    dv.error = "Odaberite jednu od ponuđenih opcija."
    dv.errorTitle = "Nevažeći status"
    dv.prompt = "Odaberite status pregleda."
    dv.promptTitle = "Status"
    ws.add_data_validation(dv)
    dv.add(f"I2:I{last_row}")

    # Conditional formatting on Status column
    status_range = f"I2:I{last_row}"
    ws.conditional_formatting.add(
        status_range,
        CellIsRule(operator="equal", formula=['"Odobreno"'], fill=_GREEN_FILL),
    )
    ws.conditional_formatting.add(
        status_range,
        CellIsRule(operator="equal", formula=['"Odbijeno"'], fill=_RED_FILL),
    )
    ws.conditional_formatting.add(
        status_range,
        CellIsRule(operator="equal", formula=['"Preskočeno"'], fill=_BLUE_FILL),
    )
    ws.conditional_formatting.add(
        status_range,
        CellIsRule(operator="equal", formula=['"Za raspravu"'], fill=_YELLOW_FILL),
    )

    # Freeze panes and auto-filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:K{last_row}"

    # Sheet protection — allow sort/filter
    ws.protection.sheet = True
    ws.protection.password = "procudo"
    ws.protection.sort = False
    ws.protection.autoFilter = False


# ── Sheet 2: Cijene ──────────────────────────────────────────────────────────


def _build_sheet2(
    ws: Worksheet,
    extractions: list[ClientExtraction],
    *,
    hrk_rate: float = 7.53450,
) -> None:
    """Build 'Cijene' (Pricing) sheet."""
    ws.title = "Cijene"

    headers = [
        "Klijent",
        "Usluga",
        "Trenutna cijena",
        "Valuta",
        "EUR protuvrijednost",
        "Jedinica",
        "Nova cijena EUR",
        "% promjene",
        "Primjena od",
    ]
    widths = [25, 40, 18, 10, 20, 15, 18, 14, 15]
    _style_header(ws, headers, widths)

    sorted_ext = sorted(extractions, key=lambda e: e.folder_name.lower())

    row_idx = 2
    for ce in sorted_ext:
        ex = ce.extraction
        if not ex or not ex.pricing_items:
            continue

        client_name = ex.client_name if ex.client_name else ce.folder_name

        for item in ex.pricing_items:
            # A: Client
            _style_cell(ws, row_idx, 1, client_name)
            # B: Service
            _style_cell(ws, row_idx, 2, item.service_name)
            # C: Current price (numeric)
            _style_cell(ws, row_idx, 3, item.price_value)
            # Format price cells
            ws.cell(row=row_idx, column=3).number_format = '#,##0.00'
            # D: Currency
            _style_cell(ws, row_idx, 4, item.currency.value)
            # E: EUR equivalent (formula)
            eur_formula = (
                f'=IF(D{row_idx}="HRK",C{row_idx}/{hrk_rate},C{row_idx})'
            )
            _style_cell(ws, row_idx, 5, eur_formula)
            ws.cell(row=row_idx, column=5).number_format = '#,##0.00'
            # F: Unit
            unit = item.unit or item.designation or ""
            _style_cell(ws, row_idx, 6, unit)

            # G: New price EUR (editable)
            _style_cell(ws, row_idx, 7, None, locked=False)
            ws.cell(row=row_idx, column=7).number_format = '#,##0.00'
            # H: % change (formula, editable column for overrides)
            pct_formula = (
                f'=IF(AND(G{row_idx}<>"",E{row_idx}>0),'
                f'(G{row_idx}-E{row_idx})/E{row_idx},"")'
            )
            _style_cell(ws, row_idx, 8, pct_formula, locked=False)
            ws.cell(row=row_idx, column=8).number_format = '0.00%'
            # I: Effective date (editable)
            _style_cell(ws, row_idx, 9, None, locked=False)
            ws.cell(row=row_idx, column=9).number_format = 'DD.MM.YYYY'

            row_idx += 1

    last_row = max(row_idx - 1, 1)

    # Freeze panes and auto-filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{last_row}"

    # Sheet protection
    ws.protection.sheet = True
    ws.protection.password = "procudo"
    ws.protection.sort = False
    ws.protection.autoFilter = False


# ── Sheet 3: Inventar ────────────────────────────────────────────────────────


def _build_sheet3(
    ws: Worksheet,
    inventory: Inventory,
) -> None:
    """Build 'Inventar' (File Inventory) sheet — read-only reference."""
    ws.title = "Inventar"

    headers = [
        "Klijent",
        "Datoteka",
        "Ekstenzija",
        "Veličina (KB)",
        "Datum izmjene",
        "Klasifikacija",
        "Status",
    ]
    widths = [25, 45, 12, 14, 18, 22, 20]
    _style_header(ws, headers, widths)

    row_idx = 2
    for client in sorted(inventory.clients, key=lambda c: c.folder_name.lower()):
        for f in client.files:
            _style_cell(ws, row_idx, 1, client.folder_name)
            _style_cell(ws, row_idx, 2, f.filename)
            _style_cell(ws, row_idx, 3, f.extension)
            size_kb = round(f.size_bytes / 1024, 1) if f.size_bytes else 0
            _style_cell(ws, row_idx, 4, size_kb)
            ws.cell(row=row_idx, column=4).number_format = '#,##0.0'
            mod_date = f.modified_date.strftime("%d.%m.%Y %H:%M") if f.modified_date else ""
            _style_cell(ws, row_idx, 5, mod_date)
            _style_cell(ws, row_idx, 6, f.doc_type.value)
            _style_cell(ws, row_idx, 7, f.status.value)
            row_idx += 1

    last_row = max(row_idx - 1, 1)

    # Freeze panes and auto-filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{last_row}"

    # Full sheet protection (all locked)
    ws.protection.sheet = True
    ws.protection.password = "procudo"
    ws.protection.sort = False
    ws.protection.autoFilter = False


# ── Main entry point ─────────────────────────────────────────────────────────


def generate_spreadsheet(
    extractions: list[ClientExtraction],
    inventory: Inventory,
    config: PipelineConfig,
) -> Path:
    """Generate the control spreadsheet with 3 sheets.

    Returns:
        Path to the generated .xlsx file.
    """
    wb = Workbook()

    # Sheet 1: Client overview (uses the default sheet)
    ws1 = wb.active
    _build_sheet1(ws1, extractions, inventory)

    # Sheet 2: Pricing
    ws2 = wb.create_sheet()
    _build_sheet2(ws2, extractions, hrk_rate=config.currency.hrk_to_eur_rate)

    # Sheet 3: File inventory
    ws3 = wb.create_sheet()
    _build_sheet3(ws3, inventory)

    # Save
    output_path = config.spreadsheet_path
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))

    return output_path
