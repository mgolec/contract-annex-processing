"""Phase 3: Generate annex .docx documents from approved spreadsheet rows."""

from __future__ import annotations

import logging
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

from openpyxl import load_workbook
from rich.table import Table

from doc_pipeline.config import PipelineConfig
from doc_pipeline.models import ClientExtraction, Currency, PricingItem
from doc_pipeline.utils.croatian import hr_date, hr_number, nfc
from doc_pipeline.utils.progress import console

logger = logging.getLogger(__name__)

# HRK → EUR conversion rate — no longer hardcoded; sourced from config.currency.hrk_to_eur_rate


# ── Data classes for spreadsheet read-back ──────────────────────────────────


@dataclass
class NewPrice:
    """A single new price entry from Sheet 2."""

    service_name: str
    new_price_eur: Decimal
    effective_date: date | None = None


@dataclass
class ApprovedClient:
    """An approved client from Sheet 1 with new prices from Sheet 2."""

    client_name: str
    folder_name: str
    new_prices: list[NewPrice] = field(default_factory=list)


# ── Spreadsheet read-back ───────────────────────────────────────────────────


def _parse_date_cell(value) -> date | None:
    """Parse a date from an openpyxl cell value."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    # Try parsing string formats
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%d.%m.%Y", "%d.%m.%Y.", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def read_approved_clients(
    spreadsheet_path: Path,
    config: PipelineConfig | None = None,
) -> list[ApprovedClient]:
    """Read approved clients and their new prices from the control spreadsheet.

    Sheet 1 "Pregled klijenata":
      Col A = client name, Col B = folder name, Col I = status
      Filter: status == "Odobreno"

    Sheet 2 "Cijene":
      Col A = client name, Col B = service name,
      Col G = new price EUR (direct), Col H = % increase/decrease,
      Col J = effective date.
      If H is filled, new price = EUR_equiv * (1 + H/100).
      If only G is filled, new price = G. H takes precedence.

    Args:
        spreadsheet_path: Path to the control spreadsheet.
        config: Optional pipeline config for HRK rate fallback.
    """
    # C7: data_only=True means Excel formula results are read as cached values.
    # If the file was saved by openpyxl (not Excel), formula cells will return None.
    # We handle this below for the EUR equivalent column (E) by computing in Python.
    wb = load_workbook(str(spreadsheet_path), data_only=True)

    try:
        # ── Sheet 1: find approved clients ──────────────────────────────
        ws1 = wb["Pregled klijenata"]

        # C2: Validate Sheet 1 headers match expected structure
        EXPECTED_HEADERS_S1 = {
            1: "Klijent",
            2: "Mapa",
            9: "Status",
        }
        for col, expected in EXPECTED_HEADERS_S1.items():
            actual = ws1.cell(1, col).value
            if actual != expected:
                raise ValueError(
                    f"Neočekivano zaglavlje u stupcu {col} (Sheet 1): '{actual}' "
                    f"(očekivano: '{expected}')\n"
                    f"Unexpected header in column {col} (Sheet 1): '{actual}' "
                    f"(expected: '{expected}')\n"
                    f"The spreadsheet may have been modified. Please check the column structure."
                )

        approved: dict[str, ApprovedClient] = {}

        for row in ws1.iter_rows(min_row=2, values_only=False):
            client_name = row[0].value  # Col A
            folder_name = row[1].value  # Col B
            status = row[8].value       # Col I

            if not folder_name or not status:
                continue
            if str(status).strip() != "Odobreno":
                continue

            approved[str(folder_name).strip()] = ApprovedClient(
                client_name=str(client_name or folder_name).strip(),
                folder_name=str(folder_name).strip(),
            )

        if not approved:
            return []

        # ── Sheet 2: collect new prices for approved clients ────────────
        ws2 = wb["Cijene"]

        # C2: Validate Sheet 2 headers match expected structure
        EXPECTED_HEADERS_S2 = {
            1: "Klijent",
            2: "Usluga",
            7: "Nova cijena EUR",
            8: "% povećanja",
        }
        for col, expected in EXPECTED_HEADERS_S2.items():
            actual = ws2.cell(1, col).value
            if actual != expected:
                raise ValueError(
                    f"Neočekivano zaglavlje u stupcu {col} (Sheet 2): '{actual}' "
                    f"(očekivano: '{expected}')\n"
                    f"Unexpected header in column {col} (Sheet 2): '{actual}' "
                    f"(expected: '{expected}')\n"
                    f"The spreadsheet may have been modified. Please check the column structure."
                )

        # H16: Build mapping from client name → folder name for matching.
        # Use NFC normalization for consistent Croatian character comparison.
        name_to_folder: dict[str, str] = {}
        for ac in approved.values():
            normalized_name = unicodedata.normalize('NFC', ac.client_name).lower()
            name_to_folder[normalized_name] = ac.folder_name
            # Also index by folder_name for direct match
            normalized_folder = unicodedata.normalize('NFC', ac.folder_name).lower()
            name_to_folder[normalized_folder] = ac.folder_name

        for row in ws2.iter_rows(min_row=2, values_only=False):
            client_name_cell = row[0].value  # Col A
            service_name = row[1].value      # Col B
            current_price = row[2].value     # Col C: current price
            currency_cell = row[3].value     # Col D: currency
            eur_equiv = row[4].value         # Col E: EUR equivalent (formula — may be None)
            new_price = row[6].value         # Col G: direct price entry
            pct_increase = row[7].value      # Col H: % increase/decrease
            effective_date = row[9].value    # Col J

            if not client_name_cell:
                continue

            client_name_str = unicodedata.normalize('NFC', str(client_name_cell).strip())

            # H16: Match to approved client — try folder name first, then client name
            folder = None
            if client_name_str in approved:
                folder = client_name_str
            else:
                folder = name_to_folder.get(client_name_str.lower())

            if folder is None or folder not in approved:
                # H16: Warn about unmatched prices
                if new_price is not None or pct_increase is not None:
                    console.print(
                        f"  [yellow]Upozorenje: nova cijena za '{client_name_str}' "
                        f"ne odgovara nijednom odobrenom klijentu[/yellow]"
                    )
                continue

            # Skip rows without any price input (neither direct nor percentage)
            if new_price is None and pct_increase is None:
                continue

            # C7: If EUR equivalent cell is None (formula not cached by Excel),
            # compute it in Python as a fallback.
            if eur_equiv is None and current_price is not None:
                hrk_rate_val = Decimal("7.53450")
                if config is not None:
                    hrk_rate_val = Decimal(str(config.currency.hrk_to_eur_rate))
                try:
                    if currency_cell == "HRK":
                        eur_equiv = float(Decimal(str(current_price)) / hrk_rate_val)
                    else:
                        eur_equiv = float(current_price)
                except (TypeError, ValueError, ArithmeticError):
                    pass

            # Determine final price: percentage (H) takes precedence over direct (G)
            price_val: Decimal | None = None

            if pct_increase is not None:
                # Percentage input — calculate new price from EUR equivalent
                try:
                    pct = Decimal(str(pct_increase))
                except (TypeError, ValueError, InvalidOperation):
                    logger.warning(
                        "Invalid percentage value '%s' for %s / %s — skipping",
                        pct_increase, client_name_str, service_name,
                    )
                    continue

                if eur_equiv is None or eur_equiv == 0:
                    logger.warning(
                        "Cannot apply percentage for %s / %s — no EUR base price",
                        client_name_str, service_name,
                    )
                    continue

                base = Decimal(str(eur_equiv))
                price_val = (base * (1 + pct / 100)).quantize(
                    Decimal("0.01"), rounding=ROUND_HALF_UP
                )
                logger.debug(
                    "Percentage price: %s × (1 + %s%%) = %s for %s / %s",
                    base, pct, price_val, client_name_str, service_name,
                )
            elif new_price is not None:
                # Direct price input
                if isinstance(new_price, str):
                    from doc_pipeline.utils.croatian import parse_hr_number
                    price_val = parse_hr_number(new_price)
                    if price_val is None:
                        continue
                else:
                    try:
                        price_val = Decimal(str(new_price))
                    except (TypeError, ValueError, InvalidOperation):
                        continue

            if price_val is None:
                continue

            approved[folder].new_prices.append(
                NewPrice(
                    service_name=str(service_name or "").strip(),
                    new_price_eur=price_val,
                    effective_date=_parse_date_cell(effective_date),
                )
            )

        return list(approved.values())
    finally:
        # M43: Ensure workbook is always closed, even on exception
        wb.close()


# ── HRK → EUR conversion ───────────────────────────────────────────────────


def _hrk_to_eur(hrk_amount: float | Decimal, rate: Decimal) -> float:
    """Convert HRK to EUR using the provided conversion rate.

    Args:
        hrk_amount: Amount in HRK.
        rate: HRK-to-EUR conversion rate (e.g. Decimal("7.53450")).
    """
    result = Decimal(str(hrk_amount)) / rate
    return float(result.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


# ── Source document parser ──────────────────────────────────────────────────


@dataclass
class SourceDocData:
    """Data extracted directly from the source .docx (latest contract/annex)."""

    korisnik_direktor: str = ""
    korisnik_adresa: str = ""
    ukupno_sati: str = ""
    l1_sati: str = ""
    l2_sati: str = ""


def _parse_source_document(doc_path: Path) -> SourceDocData:
    """Parse a .docx to extract client details and hour fund.

    Looks for:
    - Header paragraph: "...kojeg zastupa direktor(ica) NAME (u dalj..."
      Also extracts address from the same paragraph.
    - Hour fund paragraph: "...fond sati...je NN sata mjesečno:"
    - L1/L2 lines: "NN sistem administrator sata (L1)"
    """
    from docx import Document

    data = SourceDocData()

    try:
        doc = Document(str(doc_path))
    except Exception:
        return data

    paragraphs = [nfc(p.text.strip()) for p in doc.paragraphs]

    for i, text in enumerate(paragraphs):
        if not text:
            continue

        # ── Client director + address from header paragraph ─────────
        # The first paragraph with "kojeg zastupa" before the "i" separator
        # is the client party. Normalize whitespace for matching (some docs
        # have double/triple spaces).
        text_norm = re.sub(r'\s+', ' ', text).lower()
        if not data.korisnik_direktor and "kojeg zastupa" in text_norm:
            # Skip if this is the Procudo paragraph
            if "procudo" in text_norm:
                continue

            # Director name: between "direktor(ica) " and next punctuation/bracket
            # Handle variable whitespace (some docs have double/triple spaces)
            m = re.search(r'direktor(?:ica|a)?\s+([^(,]+)', text, re.IGNORECASE)
            if m:
                # Normalize internal whitespace in extracted name
                name = re.sub(r'\s+', ' ', m.group(1)).strip().rstrip(',')
                data.korisnik_direktor = name

            # Address: try two patterns
            # Pattern 1: "Company, Address, City, OIB/MB:" (comma after company name)
            m_addr = re.search(r'^[^,]+,\s*(.+?)(?:,\s*(?:OIB|MB))', text)
            # Pattern 2: "CompanyName Street N, ZipCode City, OIB/MB:"
            # (no comma between company name and street — grab from street number)
            m_addr2 = re.search(r'(\S+\s+\d+\S*,\s*\d{5}\s+\S+)', text)
            addr_candidate = ""
            if m_addr:
                addr_candidate = m_addr.group(1).strip().rstrip(',')
            # If pattern 1 result looks like just a city (no street number),
            # prefer the pattern 2 result which includes the street
            if m_addr2 and (not addr_candidate or not re.search(r'\d+\S*,', addr_candidate)):
                addr_candidate = m_addr2.group(1).strip().rstrip(',')
            if addr_candidate:
                data.korisnik_adresa = re.sub(r'\s+', ' ', addr_candidate)

        # ── Hour fund: total hours ──────────────────────────────────
        if not data.ukupno_sati and "fond sati" in text.lower():
            # Primary: "je NN sati/sata/sat mjesečno"
            m = re.search(r'je\s+(\d+)\s+sat', text)
            if not m:
                # Fallback: "je NN" at end of line (next line starts with "sati")
                m = re.search(r'je\s+(\d+)\s*$', text)
            if m:
                data.ukupno_sati = m.group(1)

        # ── L1 hours (client/workstation hours) ───────────────────
        # Matches all Croatian terminology variants:
        #   "5 sistem administrator sata (L1)"
        #   "6 sistem administrator sati (L1 i L2)"
        #   "2 klijentska sata"  /  "6 klijentskih sati"
        if not data.l1_sati:
            text_lower = text.lower()
            m = None
            if "sistem administrator" in text_lower:
                m = re.search(r'(\d+)\s+sistem\s+administrator', text, re.IGNORECASE)
            elif "klijentsk" in text_lower:
                m = re.search(r'(\d+)\s+klijentsk\w*\s+sat', text, re.IGNORECASE)
            if m:
                data.l1_sati = m.group(1)

        # ── L2 hours (server/engineer hours) ──────────────────────
        # Matches all Croatian terminology variants:
        #   "1 sistem inženjer sat – (L2)"  /  "(L3)"
        #   "1 poslužiteljski sat"  /  "3 poslužiteljska sata"
        if not data.l2_sati:
            text_lower = text.lower()
            m = None
            if "sistem inženjer" in text_lower:
                m = re.search(r'(\d+)\s+sistem\s+inženjer', text, re.IGNORECASE)
            elif "poslužiteljsk" in text_lower:
                m = re.search(r'(\d+)\s+poslužiteljsk\w*\s+sat', text, re.IGNORECASE)
            if m:
                data.l2_sati = m.group(1)

    # ── Fallback: infer missing L1/L2 from total hours ────────────
    # Some contracts only have total hours with no L1/L2 breakdown.
    # In that case, assign all hours to L1 and set L2 to 0.
    if data.ukupno_sati and not data.l1_sati and not data.l2_sati:
        data.l1_sati = data.ukupno_sati
        data.l2_sati = "0"
        logger.debug(
            "No L1/L2 breakdown found, assigning all %s hours to L1",
            data.ukupno_sati,
        )
    elif data.ukupno_sati and data.l1_sati and not data.l2_sati:
        # L1 found but no L2 — set L2 to 0 (some contracts only have L1)
        data.l2_sati = "0"

    return data


def _find_candidate_documents(
    extraction: ClientExtraction,
    config: PipelineConfig,
) -> list[Path]:
    """Find candidate .docx files to parse for client data.

    Returns a prioritized list: extraction source first, then contracts/annexes
    found in the client folder (sorted newest first by filename heuristics).
    """
    candidates: list[Path] = []
    source_file = extraction.source_file or ""
    folder_name = extraction.folder_name

    # 1. Extraction source file
    candidate = config.data_source_path / source_file
    if candidate.exists() and candidate.suffix == ".docx":
        candidates.append(candidate)

    # .doc.docx variant (LibreOffice naming)
    docx_variant = candidate.parent / (candidate.stem + ".doc.docx")
    if docx_variant.exists() and docx_variant not in candidates:
        candidates.append(docx_variant)

    # Converted version
    if extraction.was_converted:
        for try_path in [
            config.converted_path / Path(source_file).with_suffix(".docx").name,
            (config.converted_path / source_file).with_suffix(".docx"),
        ]:
            if try_path.exists() and try_path not in candidates:
                candidates.append(try_path)

    # 2. Scan client folder for contract/annex .docx files
    client_dir = config.data_source_path / folder_name
    if client_dir.is_dir():
        contract_files = []
        for f in client_dir.rglob("*.docx"):
            name_lower = f.name.lower()
            # Prioritize files with contract/annex keywords
            if any(kw in name_lower for kw in ("ugovor", "aneks", "anex", "dodatak")):
                contract_files.append(f)
        # Sort: annexes before contracts (more recent), then by name descending
        contract_files.sort(
            key=lambda f: (
                0 if any(kw in f.name.lower() for kw in ("aneks", "anex")) else 1,
                f.name.lower(),
            ),
            reverse=True,
        )
        for cf in contract_files:
            if cf not in candidates:
                candidates.append(cf)

    return candidates


def _parse_best_source_data(
    extraction: ClientExtraction,
    config: PipelineConfig,
) -> SourceDocData:
    """Try multiple candidate documents and return the best data found.

    Tries documents in priority order, merging data from multiple sources
    if no single document has everything.
    """
    candidates = _find_candidate_documents(extraction, config)
    best = SourceDocData()

    for doc_path in candidates:
        data = _parse_source_document(doc_path)

        # Merge: fill in any missing fields from this document
        if not best.korisnik_direktor and data.korisnik_direktor:
            best.korisnik_direktor = data.korisnik_direktor
        if not best.korisnik_adresa and data.korisnik_adresa:
            best.korisnik_adresa = data.korisnik_adresa
        if not best.ukupno_sati and data.ukupno_sati:
            best.ukupno_sati = data.ukupno_sati
        if not best.l1_sati and data.l1_sati:
            best.l1_sati = data.l1_sati
        if not best.l2_sati and data.l2_sati:
            best.l2_sati = data.l2_sati

        # Stop early if we have everything
        if all([
            best.korisnik_direktor,
            best.korisnik_adresa,
            best.ukupno_sati,
            best.l1_sati,
            best.l2_sati,
        ]):
            break

    return best


# ── Template context builder ───────────────────────────────────────────────


def _match_prices(
    extraction_items: list[PricingItem],
    new_prices: list[NewPrice],
) -> list[tuple[PricingItem, NewPrice | None]]:
    """Match extraction pricing items to new prices from spreadsheet by service name.

    Uses fuzzy string matching (thefuzz) to pair items by name, falling back to
    positional order only as a last resort. This is robust against row
    reordering or minor name edits in the spreadsheet.
    """
    from thefuzz import fuzz

    matched: list[tuple[PricingItem, NewPrice | None]] = []
    unmatched_prices = list(new_prices)

    for item in extraction_items:
        best_match: NewPrice | None = None
        best_score = 0
        best_idx = -1

        for i, price in enumerate(unmatched_prices):
            if item.service_name and price.service_name:
                score = fuzz.ratio(
                    item.service_name.lower().strip(),
                    price.service_name.lower().strip(),
                )
                if score > best_score:
                    best_score = score
                    best_match = price
                    best_idx = i

        if best_match is not None and best_score >= 70:
            matched.append((item, best_match))
            unmatched_prices.pop(best_idx)
        else:
            matched.append((item, None))

    # Warn about unmatched new prices
    if unmatched_prices:
        for p in unmatched_prices:
            console.print(
                f"  [yellow]Upozorenje: nova cijena za '{p.service_name}' "
                f"nema odgovarajuću stavku[/yellow]"
            )

    return matched


def build_context(
    extraction: ClientExtraction,
    approved: ApprovedClient,
    config: PipelineConfig,
    annex_number: str,
    effective_date: date,
) -> dict:
    """Build the Jinja2 template context for a single client annex."""
    ex = extraction.extraction
    if ex is None:
        raise ValueError(f"No extraction data for {extraction.folder_name}")

    is_hrk = ex.currency == Currency.HRK
    hrk_rate = Decimal(str(config.currency.hrk_to_eur_rate))
    matched = _match_prices(ex.pricing_items, approved.new_prices)
    logger.debug("Matched prices for %s: %d items", extraction.folder_name, len(matched))

    # ── Parse source documents for director, address, hours ─────────
    src_data = _parse_best_source_data(extraction, config)

    # ── Fallback: try to recover hours from extraction notes ──────
    # The Claude extraction often captures hour info in the notes field
    # (e.g., "Monthly hour allocation: 3 hours total (2 client hours + 1 server hour).")
    if ex.notes and (not src_data.ukupno_sati or not src_data.l1_sati):
        notes_text = " ".join(ex.notes) if isinstance(ex.notes, list) else str(ex.notes)
        # Total hours: "N hours total" or "N hours monthly" or "fund of N hours"
        if not src_data.ukupno_sati:
            for pattern in [
                r'(\d+)\s+hours?\s+total',
                r'(\d+)\s+hours?\s+monthly',
                r'(?:fund|allocation)\s+(?:of\s+)?(\d+)\s+hours?',
                r'includes?\s+(\d+)\s+hours?',
                r'is\s+(\d+)\s+hours?',
            ]:
                m = re.search(pattern, notes_text, re.IGNORECASE)
                if m:
                    src_data.ukupno_sati = m.group(1)
                    logger.debug("Recovered total hours from notes: %s", m.group(1))
                    break
        # L1/client hours: "N client hours" or "N L1 hours" or "N hours for workstations"
        if not src_data.l1_sati:
            for pattern in [
                r'(\d+)\s+(?:client|L1|workstation|klijentsk\w*)\s+hours?',
                r'(\d+)\s+hours?\s+for\s+(?:workstation|radnih|client|desktop)',
            ]:
                m = re.search(pattern, notes_text, re.IGNORECASE)
                if m:
                    src_data.l1_sati = m.group(1)
                    logger.debug("Recovered L1 hours from notes: %s", m.group(1))
                    break
        # L2/server hours: "N server hours" or "N L2 hours" or "N hour for server"
        if not src_data.l2_sati:
            for pattern in [
                r'(\d+)\s+(?:server|L2|L3|poslužiteljsk\w*|engineer)\s+hours?',
                r'(\d+)\s+hours?\s+for\s+(?:server|poslužitelj)',
            ]:
                m = re.search(pattern, notes_text, re.IGNORECASE)
                if m:
                    src_data.l2_sati = m.group(1)
                    logger.debug("Recovered L2 hours from notes: %s", m.group(1))
                    break
        # Infer from total if still missing
        if src_data.ukupno_sati and not src_data.l1_sati and not src_data.l2_sati:
            src_data.l1_sati = src_data.ukupno_sati
            src_data.l2_sati = "0"
        elif src_data.ukupno_sati and src_data.l1_sati and not src_data.l2_sati:
            src_data.l2_sati = "0"

    # Build stavke (pricing table rows)
    stavke = []
    for item, new_price in matched:
        if new_price is None:
            # No new price — use old price (converted if HRK)
            old_eur = _hrk_to_eur(item.price_value, hrk_rate) if is_hrk else item.price_value
            price_str = hr_number(old_eur) if old_eur else ""
        else:
            price_str = hr_number(new_price.new_price_eur)

        stavke.append({
            "pozicija": item.position,
            "opis": item.service_name,
            "oznaka": item.designation,
            "mjera": item.unit,
            "kolicina": item.quantity,
            "cijena": price_str,
        })

    # Monthly fee = first pricing item's new price (if available)
    mjesecna_naknada = ""
    if matched:
        item0, new_price0 = matched[0]
        if new_price0 is not None:
            mjesecna_naknada = hr_number(new_price0.new_price_eur)
        elif item0.price_value:
            val = _hrk_to_eur(item0.price_value, hrk_rate) if is_hrk else item0.price_value
            mjesecna_naknada = hr_number(val)

    context = {
        # Client details — pulled from source document
        "korisnik_naziv": ex.client_name or extraction.folder_name,
        "korisnik_oib": ex.client_oib or "___________",
        "korisnik_adresa": src_data.korisnik_adresa or "___________",
        "korisnik_direktor": src_data.korisnik_direktor or "___________",
        # Procudo details
        "davatelj_naziv": config.general.company_name,
        "davatelj_oib": config.general.company_oib,
        "davatelj_adresa": config.general.company_address,
        "davatelj_direktor": config.general.company_director,
        # Document metadata — reference the latest existing document
        # If extraction source is an annex, new annex references that annex;
        # if it's a contract, new annex references the contract.
        "datum_aneksa": hr_date(effective_date),
        "broj_aneksa": annex_number,
        "referentni_broj": ex.contract_number if ex.contract_number else "___________",
        "referentni_naziv_gen": "Aneksa" if ex.document_type == "annex" else "Ugovora",
        "referentni_naziv_nom": "Aneks" if ex.document_type == "annex" else "Ugovor",
        "datum_referentnog": ex.document_date or "___________",
        # Pricing
        "mjesecna_naknada": mjesecna_naknada,
        "valuta_konverzija": is_hrk,
        "stavke": stavke,
        # Hours — pulled from source document
        "ukupno_sati": src_data.ukupno_sati or "___",
        "l1_sati": src_data.l1_sati or "___",
        "l2_sati": src_data.l2_sati or "___",
        # Static
        "vat_note": config.generation.vat_note,
        "mjesto": config.general.default_location,
    }

    # Warn about any remaining placeholder values
    _PLACEHOLDER_FIELDS = {
        "korisnik_oib": "OIB klijenta / Client OIB",
        "korisnik_adresa": "Adresa klijenta / Client address",
        "korisnik_direktor": "Direktor klijenta / Client director",
        "referentni_broj": "Broj ref. dokumenta / Reference doc number",
        "datum_referentnog": "Datum ref. dokumenta / Reference doc date",
        "ukupno_sati": "Ukupno sati / Total hours",
        "l1_sati": "L1 sati / L1 hours",
        "l2_sati": "L2 sati / L2 hours",
    }
    missing = []
    for field, label in _PLACEHOLDER_FIELDS.items():
        val = context.get(field, "")
        if not val or "___" in str(val):
            missing.append(label)
    if missing:
        logger.warning(
            "Placeholder values in annex for %s: %s",
            extraction.folder_name, ", ".join(missing),
        )

    return context


# ── Preview table ───────────────────────────────────────────────────────────


def _calc_avg_change(
    extraction: ClientExtraction,
    approved: ApprovedClient,
    hrk_rate: Decimal,
) -> str:
    """Calculate average percentage price change for preview."""
    ex = extraction.extraction
    if not ex or not ex.pricing_items or not approved.new_prices:
        return "—"

    is_hrk = ex.currency == Currency.HRK
    changes = []
    for i, item in enumerate(ex.pricing_items):
        if i >= len(approved.new_prices):
            break
        old_val = item.price_value
        if old_val is None or old_val == 0:
            continue
        if is_hrk:
            old_val = Decimal(str(_hrk_to_eur(old_val, hrk_rate)))
        new_val = approved.new_prices[i].new_price_eur
        changes.append(float((new_val - old_val) / old_val * 100))

    if not changes:
        return "—"
    avg = sum(changes) / len(changes)
    return f"{avg:+.1f}%"


def print_preview(
    generation_plan: list[tuple[ApprovedClient, ClientExtraction, str]],
    hrk_rate: Decimal,
) -> None:
    """Print a Rich preview table of what will be generated.

    Args:
        generation_plan: list of (approved, extraction, annex_number) tuples
        hrk_rate: HRK-to-EUR conversion rate from config.
    """
    table = Table(title="Annex Generation Preview", show_lines=True)
    table.add_column("#", justify="right", style="dim")
    table.add_column("Client", style="bold")
    table.add_column("Annex #")
    table.add_column("Services", justify="right")
    table.add_column("Avg % Change")
    table.add_column("HRK → EUR")

    for i, (approved, extraction, annex_num) in enumerate(generation_plan, 1):
        ex = extraction.extraction
        n_services = len(approved.new_prices)
        avg_change = _calc_avg_change(extraction, approved, hrk_rate)
        is_hrk = ex and ex.currency == Currency.HRK

        table.add_row(
            str(i),
            approved.client_name,
            annex_num,
            str(n_services),
            avg_change,
            "[yellow]Yes[/yellow]" if is_hrk else "No",
        )

    console.print(table)
    console.print(
        f"\n[bold]{len(generation_plan)}[/bold] annexes will be generated."
    )


# ── Annex numbering ──────────────────────────────────────────────────────


def _detect_next_annex_number(*scan_dirs: Path) -> int:
    """Scan directories for existing annex files and return next sequence number.

    Looks for the pattern U-YY-NN in filenames where YY matches the current year
    (e.g., Aneks_U-26-05.docx in 2026) and returns max(NN) + 1.

    Args:
        *scan_dirs: One or more directories to scan for existing annexes.
    """
    current_yy = datetime.now().strftime('%y')
    max_num = 0
    pattern = re.compile(rf'U-{current_yy}-(\d+)', re.IGNORECASE)
    for scan_dir in scan_dirs:
        if scan_dir.exists():
            for f in scan_dir.rglob("*.docx"):
                match = pattern.search(f.stem)
                if match:
                    num = int(match.group(1))
                    max_num = max(max_num, num)
    return max_num + 1


# ── Main generation logic ──────────────────────────────────────────────────


def run_generation(
    config: PipelineConfig,
    *,
    start_number: int | None = None,
    client_names: list[str] | None = None,
    dry_run: bool = False,
    force: bool = False,
    output_to_source: bool = False,
) -> list[Path]:
    """Phase 3: Generate annex documents from approved spreadsheet rows.

    Args:
        config: Pipeline configuration
        start_number: Starting sequence number for annex numbering.
            If None, auto-detects from existing files in the output directory
            and source contracts directory for the current year.
        client_names: Optional filter — only generate for these clients
        dry_run: Show preview only, don't write files
        output_to_source: If True, write annexes into the original contract
            folder (config.source_path) instead of the output annexes folder.
        force: Overwrite existing output files

    Returns:
        List of paths to generated .docx files
    """
    # ── Validate prerequisites ──────────────────────────────────────
    if not config.spreadsheet_path.exists():
        console.print(
            "[red]Control spreadsheet not found.[/red] "
            f"Expected: {config.spreadsheet_path}\n"
            "Run 'pipeline extract' first."
        )
        return []

    if not config.template_path.exists():
        console.print(
            "[red]Template not found.[/red] "
            f"Expected: {config.template_path}\n"
            "Run 'python scripts/create_template.py' first."
        )
        return []

    # ── Read approved clients from spreadsheet ──────────────────────
    console.print("[bold]Reading control spreadsheet...[/bold]")
    approved_list = read_approved_clients(config.spreadsheet_path, config=config)

    if not approved_list:
        console.print(
            "[yellow]No approved clients found.[/yellow] "
            'Mark clients as "Odobreno" in the Status column (I) of the spreadsheet.'
        )
        return []

    # ── Filter by client names if provided ──────────────────────────
    if client_names:
        # Match on folder name (case-insensitive, partial match)
        filtered = []
        for ac in approved_list:
            for name in client_names:
                if name.lower() in ac.folder_name.lower():
                    filtered.append(ac)
                    break
        if not filtered:
            console.print(
                f"[yellow]None of the specified clients ({', '.join(client_names)}) "
                f"are approved in the spreadsheet.[/yellow]"
            )
            return []
        approved_list = filtered

    # ── Load extractions for approved clients ───────────────────────
    generation_plan: list[tuple[ApprovedClient, ClientExtraction, str]] = []
    skipped: list[str] = []
    year_prefix = f"U-{datetime.now().strftime('%y')}-"

    # Note: all timestamps are local time (CET/CEST for Croatia). No timezone conversion needed.

    # M27: Auto-detect next annex number if not explicitly provided
    # Scan both output and source directories for current-year annexes
    if start_number is None:
        seq = _detect_next_annex_number(config.annexes_output_path, config.source_path)
        console.print(f"  Auto-detected next annex number: {seq}")
    else:
        seq = start_number

    # Sort alphabetically by folder name for consistent numbering
    approved_list.sort(key=lambda ac: ac.folder_name.lower())

    for ac in approved_list:
        json_path = config.extractions_path / f"{ac.folder_name}.json"
        if not json_path.exists():
            console.print(
                f"  [yellow]Skipping {ac.folder_name}: "
                f"no extraction JSON found[/yellow]"
            )
            skipped.append(ac.folder_name)
            continue

        extraction = ClientExtraction.load(json_path)
        if not extraction.extraction or not extraction.extraction.pricing_items:
            console.print(
                f"  [yellow]Skipping {ac.folder_name}: "
                f"no pricing items in extraction[/yellow]"
            )
            skipped.append(ac.folder_name)
            continue

        if not ac.new_prices:
            console.print(
                f"  [yellow]Skipping {ac.folder_name}: "
                f"no new prices in spreadsheet[/yellow]"
            )
            skipped.append(ac.folder_name)
            continue

        annex_number = f"{year_prefix}{seq:02d}"
        generation_plan.append((ac, extraction, annex_number))
        seq += 1

    if not generation_plan:
        console.print("[yellow]No clients ready for annex generation.[/yellow]")
        return []

    if skipped:
        console.print(
            f"\n[dim]Skipped {len(skipped)} clients: "
            f"{', '.join(skipped)}[/dim]"
        )

    # ── Show preview ────────────────────────────────────────────────
    hrk_rate = Decimal(str(config.currency.hrk_to_eur_rate))
    console.print()
    print_preview(generation_plan, hrk_rate)

    if dry_run:
        console.print("\n[dim]Dry run — no files generated.[/dim]")
        return []

    # ── Ask for confirmation ────────────────────────────────────────
    console.print()
    proceed = console.input("[bold]Proceed with generation? [y/N]: [/bold]")
    if proceed.strip().lower() not in ("y", "yes"):
        console.print("[dim]Cancelled.[/dim]")
        return []

    # ── Generate annexes ────────────────────────────────────────────
    from docxtpl import DocxTemplate

    generated: list[Path] = []
    default_date = _parse_date_cell(config.generation.default_effective_date) or date.today()

    for ac, extraction, annex_number in generation_plan:
        # Determine effective date: first price's date or default
        eff_date = default_date
        for np in ac.new_prices:
            if np.effective_date:
                eff_date = np.effective_date
                break

        # Build context
        context = build_context(extraction, ac, config, annex_number, eff_date)

        # M23: Validate critical context fields before rendering
        _PLACEHOLDER_PATTERNS = ["___", "________", "N/A"]
        _REQUIRED_CONTEXT_FIELDS = ["korisnik_naziv", "korisnik_oib", "referentni_broj"]
        ctx_warnings = []
        for ctx_field in _REQUIRED_CONTEXT_FIELDS:
            val = context.get(ctx_field, "")
            if not val or any(p in str(val) for p in _PLACEHOLDER_PATTERNS):
                ctx_warnings.append(f"  Missing/placeholder: {ctx_field} = '{val}'")
        if ctx_warnings:
            console.print(f"  [yellow]Upozorenje za {ac.client_name}:[/yellow]")
            for w in ctx_warnings:
                console.print(f"    [yellow]{w}[/yellow]")

        # Output path
        if output_to_source:
            out_dir = config.source_path / ac.folder_name
        else:
            out_dir = config.annexes_output_path / ac.folder_name
        out_file = out_dir / f"Aneks_{annex_number}.docx"

        if out_file.exists() and not force:
            console.print(
                f"  [yellow]Skipping {ac.folder_name}: "
                f"output already exists ({out_file.name}). Use --force to overwrite.[/yellow]"
            )
            continue

        # Render template
        tpl = DocxTemplate(str(config.template_path))
        tpl.render(context)

        out_dir.mkdir(parents=True, exist_ok=True)
        tpl.save(str(out_file))
        generated.append(out_file)
        logger.debug("Generated annex: %s", out_file)

        try:
            rel_path = out_file.relative_to(config.output_path)
        except ValueError:
            rel_path = out_file.relative_to(config.source_path)
        console.print(f"  [green]Generated:[/green] {rel_path}")

    output_location = config.source_path if output_to_source else config.annexes_output_path
    console.print(
        f"\n[bold green]Done![/bold green] "
        f"{len(generated)} annexes generated in {output_location}"
    )

    return generated


# ── Template validation ─────────────────────────────────────────────────────


REQUIRED_VARIABLES = {
    "korisnik_naziv",
    "korisnik_oib",
    "korisnik_adresa",
    "korisnik_direktor",
    "davatelj_naziv",
    "davatelj_oib",
    "davatelj_adresa",
    "davatelj_direktor",
    "datum_aneksa",
    "broj_aneksa",
    "referentni_broj",
    "referentni_naziv_gen",
    "referentni_naziv_nom",
    "datum_referentnog",
    "mjesecna_naknada",
    "valuta_konverzija",
    "stavke",
    "ukupno_sati",
    "l1_sati",
    "l2_sati",
    "vat_note",
    "mjesto",
}


def validate_template(template_path: Path) -> tuple[bool, list[str]]:
    """Validate that the template contains all required Jinja2 variables.

    Returns:
        (is_valid, list of issue messages)
    """
    from docxtpl import DocxTemplate

    if not template_path.exists():
        return False, [f"Template file not found: {template_path}"]

    tpl = DocxTemplate(str(template_path))
    found = tpl.get_undeclared_template_variables()

    issues = []
    missing = REQUIRED_VARIABLES - found
    extra = found - REQUIRED_VARIABLES

    if missing:
        issues.append(f"Missing variables: {', '.join(sorted(missing))}")
    if extra:
        # Extra variables are just informational, not an error
        issues.append(f"Extra variables (OK): {', '.join(sorted(extra))}")

    is_valid = len(missing) == 0
    return is_valid, issues
